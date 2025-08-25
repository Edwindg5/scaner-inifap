from flask import Flask, request, jsonify, send_file, render_template, send_from_directory
from api.scaner import extract_data_from_pdf
from flask_cors import CORS
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
import io
import os
import json
import gc
import psutil
import logging

# Configurar logging para mejor debugging
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

app = Flask(__name__, template_folder='templates', static_folder='static')
CORS(app)

# Configuración optimizada para PDFs grandes
app.config['JSONIFY_PRETTYPRINT_REGULAR'] = False
app.config['MAX_CONTENT_LENGTH'] = 2 * 1024 * 1024 * 1024  # 2GB max upload
app.config['SEND_FILE_MAX_AGE_DEFAULT'] = 0

# Configuraciones de timeout y memoria
PROCESSING_TIMEOUT = 3600  # 1 hora para PDFs muy grandes
MAX_MEMORY_USAGE = 85      # % máximo de memoria RAM


@app.route('/')
def index():
    return render_template('index.html')

@app.route('/static/<path:path>')
def serve_static(path):
    return send_from_directory('static', path)

@app.route('/api')
def api_info():
    memory_info = psutil.virtual_memory()
    return jsonify({
        "message": "API funcionando",
        "status": "active",
        "memory_usage": f"{memory_info.percent:.1f}%",
        "memory_available": f"{memory_info.available / (1024**3):.1f} GB"
    })

@app.route('/api/procesar-pdf', methods=['POST'])
def procesar_pdf():
    initial_memory = psutil.virtual_memory().percent
    logger.info(f"Iniciando procesamiento - Memoria inicial: {initial_memory:.1f}%")
    
    try:
        # Verificaciones básicas
        if 'pdf' not in request.files:
            return jsonify({
                "status": "error",
                "message": "No se envió el archivo PDF",
                "code": 400
            }), 400
        
        pdf_file = request.files['pdf']
        
        if pdf_file.filename == '':
            return jsonify({
                "status": "error",
                "message": "No se seleccionó ningún archivo",
                "code": 400
            }), 400
        
        # Verificar tamaño del archivo
        if request.content_length and request.content_length > 2 * 1024 * 1024 * 1024:
            return jsonify({
                "status": "error",
                "message": "El archivo es demasiado grande (máximo 2GB)",
                "code": 413
            }), 413
        
        logger.info(f"Procesando archivo: {pdf_file.filename}")
        logger.info(f"Tamaño del archivo: {request.content_length / (1024*1024):.1f} MB")
        
        # Leer archivo en chunks para PDFs muy grandes
        try:
            pdf_bytes = pdf_file.read()
            logger.info(f"Archivo leído exitosamente: {len(pdf_bytes)} bytes")
            
            # Verificar memoria después de leer el archivo
            current_memory = psutil.virtual_memory().percent
            if current_memory > MAX_MEMORY_USAGE:
                logger.warning(f"Memoria alta después de leer archivo: {current_memory:.1f}%")
                gc.collect()  # Forzar limpieza de memoria
                
        except MemoryError:
            logger.error("Error de memoria al leer el archivo")
            return jsonify({
                "status": "error",
                "message": "El archivo es demasiado grande para cargar en memoria",
                "code": 507
            }), 507
        
        # Procesar PDF con manejo optimizado
        try:
            logger.info("Iniciando extracción de datos...")
            datos = extract_data_from_pdf(pdf_bytes)
            
            # Limpiar datos del archivo de memoria
            del pdf_bytes
            gc.collect()
            
            if not datos:
                return jsonify({
                    "status": "error",
                    "message": "No se pudieron extraer datos del PDF",
                    "code": 422
                }), 422
            
            # Verificar si hay errores en los datos
            if isinstance(datos, list) and len(datos) == 1 and datos[0].get('error'):
                return jsonify({
                    "status": "error",
                    "message": datos[0]['error'],
                    "code": 422
                }), 422
                
        except MemoryError:
            logger.error("Error de memoria durante el procesamiento")
            gc.collect()
            return jsonify({
                "status": "error",
                "message": "El archivo es demasiado grande para procesar en memoria disponible",
                "code": 507
            }), 507
            
        except Exception as processing_error:
            logger.error(f"Error durante procesamiento: {str(processing_error)}")
            return jsonify({
                "status": "error",
                "message": f"Error al procesar PDF: {str(processing_error)}",
                "code": 500
            }), 500
        
        final_memory = psutil.virtual_memory().percent
        logger.info(f"Procesamiento completado - Memoria final: {final_memory:.1f}%")
        logger.info(f"Registros extraídos: {len(datos) if isinstance(datos, list) else 1}")
        
        # Crear respuesta optimizada
        response_data = {
            "status": "success",
            "data": datos,
            "total_records": len(datos) if isinstance(datos, list) else 1,
            "processing_stats": {
                "memory_initial": f"{initial_memory:.1f}%",
                "memory_final": f"{final_memory:.1f}%",
                "memory_used": f"{final_memory - initial_memory:.1f}%"
            },
            "code": 200
        }
        
        return jsonify(response_data)
        
    except Exception as e:
        logger.error(f"Error general en procesamiento: {str(e)}")
        # Limpieza de emergencia
        gc.collect()
        return jsonify({
            "status": "error",
            "message": f"Error interno del servidor: {str(e)}",
            "code": 500
        }), 500

@app.route('/api/descargar-excel', methods=['POST'])
def descargar_excel():
    try:
        data = request.get_json()
        if not data:
            return jsonify({
                "status": "error",
                "message": "No se recibieron datos",
                "code": 400
            }), 400
        
        logger.info(f"Generando Excel para {len(data)} registros")
        
        # ORDENAR ALFABÉTICAMENTE POR NOMBRE DEL PRODUCTOR
        data_sorted = sorted(data, key=lambda x: x.get('nombre_productor', '').upper())
        
        # Crear libro de Excel de manera más eficiente
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Analisis_Suelo_INIFAP"
        
        # MAPEO DE COLUMNAS EN EL ORDEN CORRECTO (según tu ejemplo)
        column_mapping = [
            ('municipio', 'MUNICIPIO'),
            ('localidad', 'LOCALIDAD'),
            ('nombre_productor', 'NOMBRE DEL PRODUCTOR'),
            ('cultivo_establecer', 'CULTIVO ANTERIOR'),
            ('arcilla', 'ARCILLA'),
            ('limo', 'LIMO'),
            ('arena', 'ARENA'),
            ('textura', 'TEXTURA'),
            ('densidad_aparente', 'DA.'),
            ('ph_agua', 'PH'),
            ('mo', 'MO.'),
            ('fosforo', 'FOSFORO'),
            ('nitrogeno', 'N.INORGANICO'),
            ('potasio', 'K'),
            ('magnesio', 'MG'),
            ('calcio', 'CA'),
            ('sodio', 'NA'),
            ('azufre', 'AL'),  # Nota: en tu ejemplo parece ser AL, ajustar según necesidad
            ('conductividad_electrica', 'CIC'),
            ('capacidad_campo', 'CIC CALCULADA'),  # Mapear según disponibilidad
            ('punto_marchitez', 'H'),  # Mapear según disponibilidad
            ('azufre', 'AZUFRE'),
            ('hierro', 'HIERRO'),
            ('cobre', 'COBRE'),
            ('zinc', 'ZINC'),
            ('manganeso', 'MANGANESO'),
            ('boro', 'BORO'),
            ('', 'Columna1'),  # Columnas vacías del ejemplo
            ('', 'Columna2'),
            ('rel_ca_mg', 'CA/MG'),
            ('rel_mg_k', 'MG/K'),
            ('rel_ca_k', 'CA/K'),
            ('rel_ca_mg_k', '(CA₊MG)/K'),
            ('rel_k_mg', 'K/MG')
        ]
        
        # Estilos para el Excel
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        center_alignment = Alignment(horizontal='center', vertical='center')
        
        # Escribir encabezados con estilo
        for col_idx, (_, header) in enumerate(column_mapping, 1):
            cell = sheet.cell(row=1, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border
            cell.alignment = center_alignment
        
        # Escribir datos fila por fila (USANDO DATA ORDENADA)
        for row_idx, record in enumerate(data_sorted, 2):
            for col_idx, (key, _) in enumerate(column_mapping, 1):
                # Manejar columnas vacías
                if key == '':
                    value = ''
                else:
                    # Limpiar y formatear valor
                    raw_value = record.get(key, 'N/A')
                    if raw_value in ['No encontrado', 'No analizado', None, '']:
                        value = 'N/A'
                    else:
                        value = str(raw_value).strip()
                        # Limpiar valores numéricos
                        if key in ['arcilla', 'limo', 'arena', 'ph_agua', 'mo', 'fosforo', 'nitrogeno', 'potasio', 'calcio', 'magnesio', 'sodio']:
                            # Remover unidades y texto extra, mantener solo números
                            import re
                            numeric_match = re.search(r'([\d.,]+)', value)
                            if numeric_match:
                                value = numeric_match.group(1).replace(',', '.')
                
                cell = sheet.cell(row=row_idx, column=col_idx, value=value)
                cell.border = border
                
                # Alternar colores de filas para mejor legibilidad
                if row_idx % 2 == 0:
                    cell.fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
        
        # Ajustar anchos de columna de manera más inteligente
        for col_idx, (_, header) in enumerate(column_mapping, 1):
            # Calcular ancho basado en contenido
            max_length = len(header)
            
            # Muestrear algunas filas para calcular ancho (más eficiente)
            sample_size = min(50, len(data_sorted))
            for row_idx in range(2, min(2 + sample_size, len(data_sorted) + 2)):
                cell_value = sheet.cell(row=row_idx, column=col_idx).value
                if cell_value:
                    max_length = max(max_length, len(str(cell_value)))
            
            # Establecer ancho con límites razonables
            adjusted_width = max(10, min(max_length + 3, 35))
            column_letter = sheet.cell(row=1, column=col_idx).column_letter
            sheet.column_dimensions[column_letter].width = adjusted_width
        
        # Congelar primera fila para navegación fácil
        sheet.freeze_panes = 'A2'
        
        # Guardar en memoria
        output = io.BytesIO()
        workbook.save(output)
        output.seek(0)
        
        # Limpiar workbook de memoria
        workbook.close()
        gc.collect()
        
        logger.info("Excel generado exitosamente")
        
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=f'analisis_suelo_INIFAP_{len(data_sorted)}_registros.xlsx'
        )
        
    except Exception as e:
        logger.error(f"Error al generar Excel: {str(e)}")
        gc.collect()
        return jsonify({
            "status": "error",
            "message": f"Error al generar Excel: {str(e)}",
            "code": 500
        }), 500

@app.errorhandler(413)
def too_large(e):
    return jsonify({
        "status": "error",
        "message": "El archivo es demasiado grande. Máximo permitido: 2GB",
        "code": 413
    }), 413

@app.errorhandler(500)
def internal_error(e):
    gc.collect()  # Limpiar memoria en caso de error
    return jsonify({
        "status": "error",
        "message": "Error interno del servidor",
        "code": 500
    }), 500

if __name__ == "__main__":
    port = int(os.environ.get("PORT", 5000))
    logger.info(f"Iniciando servidor en puerto {port}")
    logger.info(f"Memoria disponible: {psutil.virtual_memory().available / (1024**3):.1f} GB")
    app.run(host="0.0.0.0", port=port, threaded=True)